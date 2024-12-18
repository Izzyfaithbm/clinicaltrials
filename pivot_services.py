from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment

def get_first_pivot(writer, data, title, current_row):
    titleCell = writer.sheets["Pivot by Pharma"].cell(row=current_row, column=1, value=title.upper())
    titleCell.font = Font(bold=True, size=15)

    current_row += 2  

    pivot = data.pivot_table(index="Phases",
                            columns=["Start Year"], 
                            values=["Study Title", "Enrollment"], 
                            aggfunc={"Study Title": "count", "Enrollment": "sum"},
                            margins=True,
                            margins_name="Grand Total")
    # shows values under years instead of years under values
    pivot.columns = pivot.columns.swaplevel(0, 1)
    # groups all values with same years into one, avoiding duplicate years 
    pivot.sort_index(axis=1, level=0, inplace=True)

    for col in pivot.iloc[:, :-2].columns:
        if int(col[0]) < 2019: 
            pivot = pivot.drop([col], axis=1)

    pivot.to_excel(writer, sheet_name="Pivot by Pharma", index=True, startrow=current_row) 
    colour_cols(writer.book["Pivot by Pharma"])

    current_row += len(pivot) + 6
    
    return pivot, current_row

def get_second_pivot(writer, data, title, current_column, current_revenue_row, conditions, pivot_min_max):
    max_pivot_length = 0
    longest_pivot_min_col = 0
    longest_pivot_max_col = 0

    for condition in conditions:
        condition_data = data[data["Conditions (revised)"] == condition]
        condition_data = condition_data[condition_data["Start Year"].notna()]
        condition_data = condition_data[condition_data["Start Year"].astype(int) >= 2014]

        titleCell = writer.sheets["Revenue Insights"].cell(column=current_column, row=current_revenue_row, value=title.upper() + ": " + condition.upper())
        titleCell.font = Font(bold=True, size=16, underline="double")
        
        pivot = condition_data.pivot_table(index="Study Title",
                                columns=["Phases"], 
                                values=["Enrollment", "Duration (mos)"])

        # if sponsor:condition has no results
        if pivot.empty:
            writer.sheets["Revenue Insights"].cell(column=current_column, row=current_revenue_row + 2, value="No studies found")
            pivot_min_max[title] = None
            
            current_revenue_row += 9
            continue

        pivot.columns = pivot.columns.swaplevel(0, 1)
        pivot.sort_index(axis=1, level=0, inplace=True)

        # gets max pivot length for colouring mean row 
        if len(pivot.columns) > max_pivot_length:
            max_pivot_length = len(pivot.columns)
            longest_pivot_min_col = current_column + 1
            longest_pivot_max_col = current_column + len(pivot.columns) + 1
        
        # get every enrollment and duration column
        enrollment_columns = [col for col in pivot.columns if "Enrollment" in col]
        duration_columns = [col for col in pivot.columns if "Duration (mos)" in col]
        all_columns = enrollment_columns + duration_columns

        # calculate and add the mean row 
        mean_values = pivot[all_columns].mean(skipna=True)
        pivot.loc["Mean"] = mean_values.round()  
        
        pivot.to_excel(writer, sheet_name="Revenue Insights", index=True, startcol=current_column, startrow=current_revenue_row) 
        sheet = writer.sheets["Revenue Insights"]

        for row_num, row in enumerate(sheet.iter_rows()):
            for cell in row:
                # colour every other row light grey
                if row_num % 2 == 0:
                    cell.fill = PatternFill(fill_type="solid", start_color="efefef")
                # colour phase/top row dark grey
                if row_num == 0:
                    cell.fill = PatternFill(fill_type="solid", start_color="c1c1c1")

        for col_num, col in enumerate(sheet.iter_cols()):
            for cell in col:
                # wraps text 
                cell.alignment = Alignment(wrap_text=True)

                # change font
                cell.font = Font(name="IBM Plex Sans", bold=cell.font.bold, size=cell.font.size, underline=cell.font.underline)

                # colours every other border dark and light
                if col_num % 2 == 0:
                    cell.border = Border(left=Side(style="thick", color="000000"))
                else:
                    cell.border = Border(left=Side(style="thin", color="808080"))            

            col_letter = get_column_letter(col[0].column)

            # widens Study Title columns
            if any(cell.value == "Study Title" for cell in col):
                sheet.column_dimensions[col_letter].width = 90
                
                # unbolds text
                for cell in col:
                    cell.font = Font(bold=False)

                # widens sponsor Title Cell width
                sponsor_col_letter = get_column_letter(col[0].column - 1)
                sheet.column_dimensions[sponsor_col_letter].width = 33
            else:
                # widens remaining cells
                sheet.column_dimensions[col_letter].width = 15

        current_revenue_row += len(pivot) + 10

    # adds mins and maxes to dict for colouring mean rows 
    if longest_pivot_max_col != 0:
        pivot_min_max[title] = {"min": longest_pivot_min_col, "max": longest_pivot_max_col}
    else:
        pivot_min_max[title] = None
    
    current_revenue_row = 1
    current_column += max_pivot_length + 4

    return pivot, current_column, pivot_min_max, current_revenue_row

def get_third_pivot(writer, data, title, current_historical_row):
    titleCell = writer.sheets["Historical View"].cell(row=current_historical_row, column=1, value=title.upper())
    titleCell.font = Font(bold=True, size=15)

    current_historical_row += 2 

    pivot = data.pivot_table(index="Conditions (revised)",
                            columns=["Start Year"], 
                            values=["Study Title", "Enrollment"], 
                            aggfunc={"Study Title": "count", "Enrollment": "sum"},
                            margins=True,
                            margins_name="Grand Total")
    
    # shows values under years instead of years under values
    pivot.columns = pivot.columns.swaplevel(0, 1)
    # groups all values with same years into one, avoiding duplicate years 
    pivot.sort_index(axis=1, level=0, inplace=True)

    for col in pivot.iloc[:, :-2].columns:
        if int(col[0]) < 2014: 
            pivot = pivot.drop([col], axis=1)

    pivot.to_excel(writer, sheet_name="Historical View", index=True, startrow=current_historical_row) # row starts underneath current_historical_row
    colour_cols(writer.book["Historical View"])

    current_historical_row += len(pivot) + 6
    
    return pivot, current_historical_row

def colour_cols(sheet):
    for col_num, col in enumerate(sheet.iter_cols(), start=1):
        for cell in col:
            # colours every two columns light blue
            if (col_num % 4 == 2 or col_num % 4 == 3):
                cell.fill = PatternFill(fill_type="solid", start_color="d6ecf3")
            else:
                # colours remaining dark blue
                cell.fill = PatternFill(fill_type="solid", start_color="add8e6")
                
            # refills border to grey
            cell.border = Border(left=Side(style="thin", color="808080"), 
                                right=Side(style="thin", color="808080"), 
                                top=Side(style="thin", color="808080"), 
                                bottom=Side(style="thin", color="808080"))

            # change font
            cell.font = Font(name="IBM Plex Sans", bold=cell.font.bold, size=cell.font.size)

# colours mean rows in revenue insights using the sponsor's longest pivot table's min and max, ignores empty ones
def colour_mean(sheet, pivot_min_max, sponsor):
    if pivot_min_max[sponsor] is not None:
        min_col = pivot_min_max[sponsor]["min"]
        max_col = pivot_min_max[sponsor]["max"]

        for row in sheet.iter_rows(min_col=min_col, max_col=max_col):
            if any(cell.value == "Mean" for cell in row):
                for cell in row:
                    cell.fill = PatternFill(fill_type="solid", start_color="000000")
                    cell.font = Font(color="00FFFFFF", bold=True)
